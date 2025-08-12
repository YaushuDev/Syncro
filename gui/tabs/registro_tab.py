# registro_tab.py
# Ubicación: /syncro_bot/gui/tabs/registro_tab.py
"""
Pestaña de registro para Syncro Bot.
Gestiona el registro completo de ejecuciones del bot incluyendo fecha, hora,
duración, perfil, estado, usuario y mensajes de error con persistencia encriptada,
exportación a Excel y envío automático por email.
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
import json
import os
import threading
from datetime import datetime, timedelta
from cryptography.fernet import Fernet

# Importaciones para Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl no está instalado. Funcionalidad Excel deshabilitada.")
    print("Instale con: pip install openpyxl")


class RegistryManager:
    """Gestor de registros de ejecuciones con encriptación"""

    def __init__(self):
        self.config_file = "execution_registry.json"
        self.key_file = "registry.key"
        self.registry = []
        self.load_registry()

    def _get_or_create_key(self):
        """Obtiene o crea la clave de encriptación"""
        if os.path.exists(self.key_file):
            with open(self.key_file, 'rb') as f:
                return f.read()
        else:
            key = Fernet.generate_key()
            with open(self.key_file, 'wb') as f:
                f.write(key)
            return key

    def _encrypt_data(self, data):
        """Encripta los datos del registro"""
        key = self._get_or_create_key()
        fernet = Fernet(key)
        json_str = json.dumps(data, ensure_ascii=True, default=str)
        encrypted_data = fernet.encrypt(json_str.encode('utf-8'))
        return encrypted_data

    def _decrypt_data(self, encrypted_data):
        """Desencripta los datos del registro"""
        try:
            key = self._get_or_create_key()
            fernet = Fernet(key)
            decrypted_data = fernet.decrypt(encrypted_data)
            return json.loads(decrypted_data.decode('utf-8'))
        except Exception:
            return []

    def add_execution_record(self, start_time, end_time=None, profile_name="Manual",
                             status="En Ejecución", user_type="Usuario", error_message=""):
        """Añade un registro de ejecución"""

        # Si no se proporciona end_time, usar start_time (para registros en curso)
        if end_time is None:
            end_time = start_time
            duration = "En curso"
        else:
            # Calcular duración
            if isinstance(start_time, str):
                start_dt = datetime.fromisoformat(start_time)
            else:
                start_dt = start_time

            if isinstance(end_time, str):
                end_dt = datetime.fromisoformat(end_time)
            else:
                end_dt = end_time

            duration_delta = end_dt - start_dt
            duration = self._format_duration(duration_delta)

        record = {
            "id": self._generate_id(),
            "fecha": start_time.date().isoformat() if hasattr(start_time, 'date') else start_time[:10],
            "hora_inicio": start_time.time().isoformat() if hasattr(start_time, 'time') else start_time[11:19],
            "hora_fin": end_time.time().isoformat() if hasattr(end_time, 'time') else end_time[11:19],
            "perfil": profile_name,
            "duracion": duration,
            "estado": status,
            "usuario": user_type,
            "error_message": error_message,
            "timestamp_inicio": start_time.isoformat() if hasattr(start_time, 'isoformat') else start_time,
            "timestamp_fin": end_time.isoformat() if hasattr(end_time, 'isoformat') else end_time,
            "created": datetime.now().isoformat()
        }

        self.registry.append(record)
        self.save_registry()
        return record

    def update_execution_record(self, record_id, end_time, status, error_message=""):
        """Actualiza un registro de ejecución cuando termina"""
        for record in self.registry:
            if record["id"] == record_id:
                # Calcular duración
                start_dt = datetime.fromisoformat(record["timestamp_inicio"])
                end_dt = end_time if hasattr(end_time, 'isoformat') else datetime.fromisoformat(end_time)

                duration_delta = end_dt - start_dt
                duration = self._format_duration(duration_delta)

                # Actualizar registro
                record["hora_fin"] = end_dt.time().isoformat()
                record["timestamp_fin"] = end_dt.isoformat()
                record["duracion"] = duration
                record["estado"] = status
                record["error_message"] = error_message

                self.save_registry()
                return record
        return None

    def _format_duration(self, duration_delta):
        """Formatea la duración de manera legible"""
        total_seconds = int(duration_delta.total_seconds())

        if total_seconds < 60:
            return f"{total_seconds}s"
        elif total_seconds < 3600:
            minutes = total_seconds // 60
            seconds = total_seconds % 60
            return f"{minutes}m {seconds}s"
        else:
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            seconds = total_seconds % 60
            return f"{hours}h {minutes}m {seconds}s"

    def get_all_records(self, limit=None):
        """Obtiene todos los registros ordenados por fecha descendente"""
        sorted_registry = sorted(self.registry,
                                 key=lambda x: x["timestamp_inicio"],
                                 reverse=True)
        if limit:
            return sorted_registry[:limit]
        return sorted_registry

    def get_filtered_records(self, date_from=None, date_to=None, status_filter=None,
                             user_filter=None, profile_filter=None):
        """Obtiene registros filtrados"""
        filtered = self.registry.copy()

        if date_from:
            filtered = [r for r in filtered if r["fecha"] >= date_from]

        if date_to:
            filtered = [r for r in filtered if r["fecha"] <= date_to]

        if status_filter and status_filter != "Todos":
            filtered = [r for r in filtered if r["estado"] == status_filter]

        if user_filter and user_filter != "Todos":
            filtered = [r for r in filtered if r["usuario"] == user_filter]

        if profile_filter and profile_filter != "Todos":
            filtered = [r for r in filtered if r["perfil"] == profile_filter]

        return sorted(filtered, key=lambda x: x["timestamp_inicio"], reverse=True)

    def get_statistics(self):
        """Obtiene estadísticas generales"""
        total = len(self.registry)
        successful = len([r for r in self.registry if r["estado"] == "Exitoso"])
        failed = len([r for r in self.registry if r["estado"] == "Fallido"])
        in_progress = len([r for r in self.registry if r["estado"] == "En Ejecución"])
        manual_executions = len([r for r in self.registry if r["usuario"] == "Usuario"])
        system_executions = len([r for r in self.registry if r["usuario"] == "Sistema"])

        return {
            "total": total,
            "successful": successful,
            "failed": failed,
            "in_progress": in_progress,
            "manual": manual_executions,
            "system": system_executions,
            "success_rate": (successful / total * 100) if total > 0 else 0
        }

    def clear_registry(self):
        """Limpia todos los registros"""
        self.registry = []
        self.save_registry()

    def clear_old_records(self, days_to_keep=30):
        """Elimina registros antiguos"""
        cutoff_date = (datetime.now() - timedelta(days=days_to_keep)).date().isoformat()
        self.registry = [r for r in self.registry if r["fecha"] >= cutoff_date]
        self.save_registry()

    def save_registry(self):
        """Guarda el registro encriptado"""
        try:
            encrypted_data = self._encrypt_data(self.registry)
            with open(self.config_file, 'wb') as f:
                f.write(encrypted_data)
            return True
        except Exception as e:
            print(f"Error guardando registro: {e}")
            return False

    def load_registry(self):
        """Carga el registro desde archivo"""
        try:
            if not os.path.exists(self.config_file):
                self.registry = []
                return

            with open(self.config_file, 'rb') as f:
                encrypted_data = f.read()

            self.registry = self._decrypt_data(encrypted_data)
        except Exception as e:
            print(f"Error cargando registro: {e}")
            self.registry = []

    def _generate_id(self):
        """Genera un ID único para el registro"""
        return f"{datetime.now().strftime('%Y%m%d%H%M%S')}{len(self.registry)}"


class ReportGenerator:
    """Generador de reportes en Excel"""

    def __init__(self):
        self.temp_files = []

    def export_to_excel(self, records, filename=None, report_title="Reporte de Ejecuciones"):
        """Exporta registros a Excel con formato profesional"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl no está disponible. Instale con: pip install openpyxl")

        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"reporte_ejecuciones_{timestamp}.xlsx"

        try:
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte de Ejecuciones"

            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            # Título del reporte
            ws.merge_cells('A1:H1')
            title_cell = ws['A1']
            title_cell.value = report_title
            title_cell.font = Font(bold=True, size=16)
            title_cell.alignment = center_alignment

            # Fecha de generación
            ws.merge_cells('A2:H2')
            date_cell = ws['A2']
            date_cell.value = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
            date_cell.alignment = center_alignment

            # Espacio
            ws['A3'] = ""

            # Headers
            headers = ['Fecha', 'Hora Inicio', 'Hora Fin', 'Perfil', 'Duración', 'Estado', 'Usuario', 'Error']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border

            # Datos
            for row_idx, record in enumerate(records, 5):
                ws.cell(row=row_idx, column=1, value=record['fecha']).border = border
                ws.cell(row=row_idx, column=2, value=record['hora_inicio']).border = border
                ws.cell(row=row_idx, column=3, value=record['hora_fin']).border = border
                ws.cell(row=row_idx, column=4, value=record['perfil']).border = border
                ws.cell(row=row_idx, column=5, value=record['duracion']).border = border

                # Estado con color
                estado_cell = ws.cell(row=row_idx, column=6, value=record['estado'])
                estado_cell.border = border
                if record['estado'] == 'Exitoso':
                    estado_cell.fill = PatternFill(start_color="D5EDDA", end_color="D5EDDA", fill_type="solid")
                elif record['estado'] == 'Fallido':
                    estado_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                elif record['estado'] == 'En Ejecución':
                    estado_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

                ws.cell(row=row_idx, column=7, value=record['usuario']).border = border
                ws.cell(row=row_idx, column=8, value=record.get('error_message', '')).border = border

            # Ajustar ancho de columnas
            column_widths = [12, 12, 12, 15, 12, 12, 10, 30]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

            # Guardar archivo
            wb.save(filename)

            # Agregar a archivos temporales para limpiar después
            if filename.startswith("temp_"):
                self.temp_files.append(filename)

            return filename, True, f"Archivo Excel generado: {filename}"

        except Exception as e:
            return None, False, f"Error generando Excel: {str(e)}"

    def cleanup_temp_files(self):
        """Limpia archivos temporales generados"""
        for temp_file in self.temp_files:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
            except:
                pass
        self.temp_files = []


class RegistroTab:
    """Pestaña de registro para Syncro Bot"""

    def __init__(self, parent_notebook):
        self.parent = parent_notebook
        self.colors = {
            'bg_primary': '#f0f0f0',
            'bg_secondary': '#e0e0e0',
            'bg_tertiary': '#ffffff',
            'text_primary': '#333333',
            'text_secondary': '#666666',
            'border': '#cccccc',
            'accent': '#0078d4',
            'success': '#107c10',
            'warning': '#ff8c00',
            'error': '#d13438',
            'info': '#0078d4'
        }

        # Servicios
        self.registry_manager = RegistryManager()
        self.report_generator = ReportGenerator()

        # ===== INTEGRACIÓN CON EMAIL =====
        self.email_tab = None
        # ===== FIN INTEGRACIÓN =====

        # Variables
        self.widgets = {}
        self.current_filter = {}
        self.selected_record = None

        # Control de secciones colapsables
        self.expanded_section = None
        self.section_frames = {}

        self.create_tab()
        self.load_records()

    def set_email_tab(self, email_tab):
        """Establece la referencia al EmailTab para envío de reportes"""
        self.email_tab = email_tab

    def create_tab(self):
        """Crear la pestaña de registro"""
        self.frame = ttk.Frame(self.parent)
        self.parent.add(self.frame, text="Registro")
        self.create_interface()

    def create_interface(self):
        """Crea la interfaz con diseño de 2 columnas"""
        # Container principal
        main_container = tk.Frame(self.frame, bg=self.colors['bg_primary'])
        main_container.pack(fill='both', expand=True, padx=15, pady=10)

        # Configurar grid para 2 columnas con separador
        main_container.grid_columnconfigure(0, weight=0, minsize=450)  # Columna izquierda
        main_container.grid_columnconfigure(1, weight=0, minsize=1)  # Separador
        main_container.grid_columnconfigure(2, weight=1, minsize=650)  # Columna derecha
        main_container.grid_rowconfigure(0, weight=1)

        # Columna izquierda - Filtros y estadísticas
        left_column = tk.Frame(main_container, bg=self.colors['bg_primary'], width=450)
        left_column.grid(row=0, column=0, sticky='ns', padx=(0, 5))
        left_column.grid_propagate(False)

        # Separador visual
        separator = tk.Frame(main_container, bg=self.colors['border'], width=1)
        separator.grid(row=0, column=1, sticky='ns', padx=5)

        # Columna derecha - Tabla de registros
        right_column = tk.Frame(main_container, bg=self.colors['bg_primary'])
        right_column.grid(row=0, column=2, sticky='nsew', padx=(5, 0))

        # Crear contenido
        self._create_left_column_collapsible(left_column)
        self._create_right_column(right_column)

    def _create_left_column_collapsible(self, parent):
        """Crea la columna izquierda con secciones colapsables"""
        parent.grid_rowconfigure(0, weight=0)  # Estadísticas
        parent.grid_rowconfigure(1, weight=0)  # Filtros
        parent.grid_rowconfigure(2, weight=0)  # Reportes ===== NUEVO =====
        parent.grid_rowconfigure(3, weight=0)  # Acciones
        parent.grid_rowconfigure(4, weight=1)  # Espaciador
        parent.grid_columnconfigure(0, weight=1)

        # Sección de estadísticas
        self._create_collapsible_section(
            parent, "stats", "📊 Estadísticas Generales",
            self._create_stats_content, row=0, default_expanded=True,
            min_height=200
        )

        # Sección de filtros
        self._create_collapsible_section(
            parent, "filters", "🔍 Filtros de Búsqueda",
            self._create_filters_content, row=1, default_expanded=False,
            min_height=280
        )

        # ===== NUEVA SECCIÓN DE REPORTES =====
        self._create_collapsible_section(
            parent, "reports", "📧 Generación de Reportes",
            self._create_reports_content, row=2, default_expanded=False,
            min_height=220
        )
        # ===== FIN NUEVA SECCIÓN =====

        # Sección de acciones
        self._create_collapsible_section(
            parent, "actions", "🎮 Acciones de Gestión",
            self._create_actions_content, row=3, default_expanded=False,
            min_height=180
        )

        # Espaciador
        spacer = tk.Frame(parent, bg=self.colors['bg_primary'])
        spacer.grid(row=4, column=0, sticky='nsew')

    def _create_right_column(self, parent):
        """Crea el contenido de la columna derecha"""
        parent.grid_rowconfigure(0, weight=0)  # Header
        parent.grid_rowconfigure(1, weight=1)  # Tabla
        parent.grid_rowconfigure(2, weight=0)  # Detalle
        parent.grid_columnconfigure(0, weight=1)

        # Header con título y controles
        header_container = tk.Frame(parent, bg=self.colors['bg_primary'])
        header_container.grid(row=0, column=0, sticky='ew', pady=(0, 10))
        self._create_records_header(header_container)

        # Tabla de registros
        table_container = tk.Frame(parent, bg=self.colors['bg_primary'])
        table_container.grid(row=1, column=0, sticky='nsew', pady=(0, 10))
        self._create_records_table(table_container)

        # Panel de detalle
        detail_container = tk.Frame(parent, bg=self.colors['bg_primary'])
        detail_container.grid(row=2, column=0, sticky='ew')
        self._create_detail_panel(detail_container)

    def _create_collapsible_section(self, parent, section_id, title, content_creator,
                                    row, default_expanded=False, min_height=150):
        """Crea una sección colapsable tipo acordeón"""
        # Container principal
        section_container = tk.Frame(parent, bg=self.colors['bg_primary'])
        section_container.configure(height=55)  # Solo header cuando está colapsada
        section_container.grid(row=row, column=0, sticky='ew', pady=(0, 10))
        section_container.grid_columnconfigure(0, weight=1)
        section_container.grid_propagate(False)

        # Frame de la tarjeta
        card = tk.Frame(section_container, bg=self.colors['bg_primary'],
                        relief='solid', bd=1)
        card.configure(highlightbackground=self.colors['border'],
                       highlightcolor=self.colors['border'],
                       highlightthickness=1)
        card.grid(row=0, column=0, sticky='ew')
        card.grid_columnconfigure(0, weight=1)

        # Header clickeable
        header = tk.Frame(card, bg=self.colors['bg_secondary'], height=45, cursor='hand2')
        header.grid(row=0, column=0, sticky='ew')
        header.grid_propagate(False)
        header.grid_columnconfigure(0, weight=1)

        # Contenido del header
        header_content = tk.Frame(header, bg=self.colors['bg_secondary'])
        header_content.grid(row=0, column=0, sticky='ew', padx=15, pady=12)
        header_content.grid_columnconfigure(0, weight=1)

        # Título
        title_label = tk.Label(header_content, text=title, bg=self.colors['bg_secondary'],
                               fg=self.colors['text_primary'], font=('Arial', 12, 'bold'),
                               cursor='hand2')
        title_label.grid(row=0, column=0, sticky='w')

        # Flecha indicadora
        arrow_label = tk.Label(header_content, text="▶",
                               bg=self.colors['bg_secondary'], fg=self.colors['accent'],
                               font=('Arial', 10, 'bold'), cursor='hand2')
        arrow_label.grid(row=0, column=1, sticky='e')

        # Content area
        content_frame = tk.Frame(card, bg=self.colors['bg_primary'])
        content_frame.grid_columnconfigure(0, weight=1)

        # Crear contenido específico
        content_creator(content_frame)

        # Guardar referencias
        self.section_frames[section_id] = {
            'container': section_container,
            'header': header,
            'content': content_frame,
            'arrow': arrow_label,
            'expanded': False,
            'min_height': min_height
        }

        # Bind eventos
        def toggle_section(event=None):
            self._toggle_section(section_id)

        header.bind("<Button-1>", toggle_section)
        title_label.bind("<Button-1>", toggle_section)
        arrow_label.bind("<Button-1>", toggle_section)

        # Expandir por defecto si se especifica
        if default_expanded:
            self._toggle_section(section_id)

    def _toggle_section(self, section_id):
        """Alterna la visibilidad de una sección"""
        current_section = self.section_frames[section_id]

        if current_section['expanded']:
            # Colapsar sección actual
            current_section['content'].grid_remove()
            current_section['arrow'].configure(text="▶")
            current_section['expanded'] = False
            current_section['container'].configure(height=55)
            current_section['container'].grid_propagate(False)
            self.expanded_section = None
        else:
            # Colapsar otra sección si está expandida
            if self.expanded_section and self.expanded_section in self.section_frames:
                expanded_section = self.section_frames[self.expanded_section]
                expanded_section['content'].grid_remove()
                expanded_section['arrow'].configure(text="▶")
                expanded_section['expanded'] = False
                expanded_section['container'].configure(height=55)
                expanded_section['container'].grid_propagate(False)

            # Expandir nueva sección
            current_section['content'].grid(row=1, column=0, sticky='ew')
            current_section['arrow'].configure(text="▼")
            current_section['expanded'] = True
            current_section['container'].configure(height=current_section['min_height'])
            current_section['container'].grid_propagate(True)
            self.expanded_section = section_id

        self.frame.update_idletasks()

    def _create_stats_content(self, parent):
        """Crea el contenido de estadísticas"""
        content = tk.Frame(parent, bg=self.colors['bg_primary'])
        content.pack(fill='x', padx=18, pady=15)

        # Grid de estadísticas
        stats_grid = tk.Frame(content, bg=self.colors['bg_primary'])
        stats_grid.pack(fill='x')

        # Fila 1: Total y Exitosos
        row1 = tk.Frame(stats_grid, bg=self.colors['bg_primary'])
        row1.pack(fill='x', pady=(0, 8))

        self._create_stat_box(row1, "📈 Total Ejecuciones:", "total_stat", "0", side='left')
        self._create_stat_box(row1, "✅ Exitosas:", "success_stat", "0", side='right')

        # Fila 2: Fallidas y En Progreso
        row2 = tk.Frame(stats_grid, bg=self.colors['bg_primary'])
        row2.pack(fill='x', pady=(0, 8))

        self._create_stat_box(row2, "❌ Fallidas:", "failed_stat", "0", side='left')
        self._create_stat_box(row2, "⏳ En Progreso:", "progress_stat", "0", side='right')

        # Fila 3: Usuario vs Sistema
        row3 = tk.Frame(stats_grid, bg=self.colors['bg_primary'])
        row3.pack(fill='x', pady=(0, 8))

        self._create_stat_box(row3, "👤 Usuario:", "manual_stat", "0", side='left')
        self._create_stat_box(row3, "🤖 Sistema:", "system_stat", "0", side='right')

        # Tasa de éxito
        success_rate_frame = tk.Frame(content, bg=self.colors['bg_tertiary'])
        success_rate_frame.pack(fill='x', pady=(15, 0))

        tk.Label(success_rate_frame, text="📊 Tasa de Éxito:", bg=self.colors['bg_tertiary'],
                 fg=self.colors['text_primary'], font=('Arial', 10)).pack(
            side='left', padx=10, pady=8)

        self.widgets['success_rate_stat'] = tk.Label(
            success_rate_frame, text="0%", bg=self.colors['bg_tertiary'],
            fg=self.colors['success'], font=('Arial', 10, 'bold')
        )
        self.widgets['success_rate_stat'].pack(side='right', padx=10, pady=8)

    def _create_stat_box(self, parent, label_text, widget_key, initial_value, side='left'):
        """Crea una caja de estadística"""
        frame = tk.Frame(parent, bg=self.colors['bg_tertiary'])
        frame.pack(side=side, fill='x', expand=True, padx=(0, 4) if side == 'left' else (4, 0))

        tk.Label(frame, text=label_text, bg=self.colors['bg_tertiary'],
                 fg=self.colors['text_primary'], font=('Arial', 9)).pack(
            side='left', padx=8, pady=6)

        self.widgets[widget_key] = tk.Label(
            frame, text=initial_value, bg=self.colors['bg_tertiary'],
            fg=self.colors['info'], font=('Arial', 9, 'bold')
        )
        self.widgets[widget_key].pack(side='right', padx=8, pady=6)

    def _create_filters_content(self, parent):
        """Crea el contenido de filtros"""
        content = tk.Frame(parent, bg=self.colors['bg_primary'])
        content.pack(fill='x', padx=18, pady=15)

        # Filtro por fechas
        date_frame = tk.Frame(content, bg=self.colors['bg_primary'])
        date_frame.pack(fill='x', pady=(0, 15))

        tk.Label(date_frame, text="📅 Rango de Fechas:", bg=self.colors['bg_primary'],
                 fg=self.colors['text_primary'], font=('Arial', 10, 'bold')).pack(anchor='w', pady=(0, 5))

        date_inputs = tk.Frame(date_frame, bg=self.colors['bg_tertiary'])
        date_inputs.pack(fill='x')

        date_inner = tk.Frame(date_inputs, bg=self.colors['bg_tertiary'])
        date_inner.pack(padx=10, pady=10)

        tk.Label(date_inner, text="Desde:", bg=self.colors['bg_tertiary'],
                 fg=self.colors['text_primary'], font=('Arial', 9)).grid(row=0, column=0, sticky='w', pady=(0, 5))

        self.widgets['date_from'] = self._create_styled_entry(date_inner)
        self.widgets['date_from'].grid(row=0, column=1, sticky='ew', padx=(5, 0), pady=(0, 5))

        tk.Label(date_inner, text="Hasta:", bg=self.colors['bg_tertiary'],
                 fg=self.colors['text_primary'], font=('Arial', 9)).grid(row=1, column=0, sticky='w')

        self.widgets['date_to'] = self._create_styled_entry(date_inner)
        self.widgets['date_to'].grid(row=1, column=1, sticky='ew', padx=(5, 0))

        date_inner.grid_columnconfigure(1, weight=1)

        # Filtros por estado, usuario, perfil
        filters_frame = tk.Frame(content, bg=self.colors['bg_primary'])
        filters_frame.pack(fill='x', pady=(0, 15))

        tk.Label(filters_frame, text="🔍 Filtros Rápidos:", bg=self.colors['bg_primary'],
                 fg=self.colors['text_primary'], font=('Arial', 10, 'bold')).pack(anchor='w', pady=(0, 5))

        filters_grid = tk.Frame(filters_frame, bg=self.colors['bg_tertiary'])
        filters_grid.pack(fill='x')

        filters_inner = tk.Frame(filters_grid, bg=self.colors['bg_tertiary'])
        filters_inner.pack(padx=10, pady=10)

        # Estado
        tk.Label(filters_inner, text="Estado:", bg=self.colors['bg_tertiary'],
                 fg=self.colors['text_primary'], font=('Arial', 9)).grid(row=0, column=0, sticky='w', pady=(0, 5))

        self.widgets['status_filter'] = ttk.Combobox(filters_inner, values=[
            "Todos", "Exitoso", "Fallido", "En Ejecución"
        ], state="readonly", width=18)
        self.widgets['status_filter'].set("Todos")
        self.widgets['status_filter'].grid(row=0, column=1, sticky='ew', padx=(5, 0), pady=(0, 5))

        # Usuario
        tk.Label(filters_inner, text="Usuario:", bg=self.colors['bg_tertiary'],
                 fg=self.colors['text_primary'], font=('Arial', 9)).grid(row=1, column=0, sticky='w', pady=(0, 5))

        self.widgets['user_filter'] = ttk.Combobox(filters_inner, values=[
            "Todos", "Usuario", "Sistema"
        ], state="readonly", width=18)
        self.widgets['user_filter'].set("Todos")
        self.widgets['user_filter'].grid(row=1, column=1, sticky='ew', padx=(5, 0), pady=(0, 5))

        # Perfil
        tk.Label(filters_inner, text="Perfil:", bg=self.colors['bg_tertiary'],
                 fg=self.colors['text_primary'], font=('Arial', 9)).grid(row=2, column=0, sticky='w')

        self.widgets['profile_filter'] = ttk.Combobox(filters_inner, values=["Todos"], state="readonly", width=18)
        self.widgets['profile_filter'].set("Todos")
        self.widgets['profile_filter'].grid(row=2, column=1, sticky='ew', padx=(5, 0))

        filters_inner.grid_columnconfigure(1, weight=1)

        # Botones de filtro
        filter_buttons = tk.Frame(content, bg=self.colors['bg_primary'])
        filter_buttons.pack(fill='x')

        self.widgets['apply_filters_btn'] = self._create_styled_button(
            filter_buttons, "🔍 Aplicar Filtros",
            self._apply_filters, self.colors['info']
        )
        self.widgets['apply_filters_btn'].pack(side='left', fill='x', expand=True, padx=(0, 5))

        self.widgets['clear_filters_btn'] = self._create_styled_button(
            filter_buttons, "🧹 Limpiar",
            self._clear_filters, self.colors['text_secondary']
        )
        self.widgets['clear_filters_btn'].pack(side='right', fill='x', expand=True, padx=(5, 0))

    # ===== NUEVA SECCIÓN DE REPORTES =====
    def _create_reports_content(self, parent):
        """Crea el contenido de generación de reportes"""
        content = tk.Frame(parent, bg=self.colors['bg_primary'])
        content.pack(fill='x', padx=18, pady=15)

        # Tipo de reporte
        report_type_frame = tk.Frame(content, bg=self.colors['bg_primary'])
        report_type_frame.pack(fill='x', pady=(0, 15))

        tk.Label(report_type_frame, text="📋 Tipo de Reporte:", bg=self.colors['bg_primary'],
                 fg=self.colors['text_primary'], font=('Arial', 10, 'bold')).pack(anchor='w', pady=(0, 5))

        self.widgets['report_type'] = ttk.Combobox(report_type_frame, values=[
            "Registros Actuales (Filtrados)",
            "Últimos 7 días",
            "Últimos 30 días",
            "Solo Exitosos",
            "Solo Fallidos",
            "Solo Ejecuciones Manuales",
            "Solo Ejecuciones Automáticas",
            "Todos los Registros"
        ], state="readonly", width=35)
        self.widgets['report_type'].set("Registros Actuales (Filtrados)")
        self.widgets['report_type'].pack(fill='x')

        # Estado del email
        email_status_frame = tk.Frame(content, bg=self.colors['bg_tertiary'])
        email_status_frame.pack(fill='x', pady=(0, 15))

        tk.Label(email_status_frame, text="📧 Email:", bg=self.colors['bg_tertiary'],
                 fg=self.colors['text_primary'], font=('Arial', 10)).pack(
            side='left', padx=10, pady=8)

        self.widgets['email_status'] = tk.Label(
            email_status_frame, text="No configurado", bg=self.colors['bg_tertiary'],
            fg=self.colors['text_secondary'], font=('Arial', 10, 'bold')
        )
        self.widgets['email_status'].pack(side='right', padx=10, pady=8)

        # Botones de reportes
        report_buttons = tk.Frame(content, bg=self.colors['bg_primary'])
        report_buttons.pack(fill='x')

        # Botón exportar Excel
        self.widgets['export_excel_btn'] = self._create_styled_button(
            report_buttons, "📊 Exportar a Excel",
            self._export_to_excel, self.colors['info']
        )
        self.widgets['export_excel_btn'].pack(fill='x', pady=(0, 10))

        # Botón enviar por email
        self.widgets['send_email_btn'] = self._create_styled_button(
            report_buttons, "📧 Generar y Enviar por Email",
            self._send_report_email, self.colors['success']
        )
        self.widgets['send_email_btn'].pack(fill='x')

        # Deshabilitar si no hay Excel o Email
        if not EXCEL_AVAILABLE:
            self.widgets['export_excel_btn'].configure(state='disabled',
                                                       text='📊 Excel no disponible (instalar openpyxl)')
            self.widgets['send_email_btn'].configure(state='disabled')

    # ===== FIN NUEVA SECCIÓN =====

    def _create_actions_content(self, parent):
        """Crea el contenido de acciones"""
        content = tk.Frame(parent, bg=self.colors['bg_primary'])
        content.pack(fill='x', padx=18, pady=15)

        # Botón refrescar
        self.widgets['refresh_btn'] = self._create_styled_button(
            content, "🔄 Refrescar Registros",
            self.load_records, self.colors['info']
        )
        self.widgets['refresh_btn'].pack(fill='x', pady=(0, 10))

        # Botón limpiar antiguos
        self.widgets['clean_old_btn'] = self._create_styled_button(
            content, "🧹 Limpiar Antiguos (30+ días)",
            self._clean_old_records, self.colors['warning']
        )
        self.widgets['clean_old_btn'].pack(fill='x', pady=(0, 10))

        # Botón limpiar todos
        self.widgets['clear_all_btn'] = self._create_styled_button(
            content, "🗑️ Eliminar Todos los Registros",
            self._clear_all_records, self.colors['error']
        )
        self.widgets['clear_all_btn'].pack(fill='x')

    def _create_records_header(self, parent):
        """Crea el header de la sección de registros"""
        header_frame = tk.Frame(parent, bg=self.colors['bg_secondary'], height=45)
        header_frame.pack(fill='x')
        header_frame.pack_propagate(False)

        # Título
        tk.Label(header_frame, text="📋 Historial de Ejecuciones",
                 bg=self.colors['bg_secondary'], fg=self.colors['text_primary'],
                 font=('Arial', 12, 'bold')).pack(side='left', padx=15, pady=12)

        # Contador de registros
        self.widgets['record_count'] = tk.Label(
            header_frame, text="0 registros", bg=self.colors['bg_secondary'],
            fg=self.colors['text_secondary'], font=('Arial', 10)
        )
        self.widgets['record_count'].pack(side='right', padx=15, pady=12)

    def _create_records_table(self, parent):
        """Crea la tabla de registros"""
        # Frame con border
        table_frame = tk.Frame(parent, bg=self.colors['border'], bd=1, relief='solid')
        table_frame.pack(fill='both', expand=True)

        # Treeview para mostrar registros
        columns = ('fecha', 'inicio', 'fin', 'perfil', 'duracion', 'estado', 'usuario')
        self.widgets['records_tree'] = ttk.Treeview(table_frame, columns=columns, show='headings')

        # Configurar columnas
        self.widgets['records_tree'].heading('fecha', text='Fecha')
        self.widgets['records_tree'].heading('inicio', text='H. Inicio')
        self.widgets['records_tree'].heading('fin', text='H. Fin')
        self.widgets['records_tree'].heading('perfil', text='Perfil')
        self.widgets['records_tree'].heading('duracion', text='Duración')
        self.widgets['records_tree'].heading('estado', text='Estado')
        self.widgets['records_tree'].heading('usuario', text='Usuario')

        self.widgets['records_tree'].column('fecha', width=90)
        self.widgets['records_tree'].column('inicio', width=70)
        self.widgets['records_tree'].column('fin', width=70)
        self.widgets['records_tree'].column('perfil', width=100)
        self.widgets['records_tree'].column('duracion', width=80)
        self.widgets['records_tree'].column('estado', width=80)
        self.widgets['records_tree'].column('usuario', width=70)

        # Scrollbars
        v_scrollbar = ttk.Scrollbar(table_frame, orient='vertical',
                                    command=self.widgets['records_tree'].yview)
        h_scrollbar = ttk.Scrollbar(table_frame, orient='horizontal',
                                    command=self.widgets['records_tree'].xview)

        self.widgets['records_tree'].configure(yscrollcommand=v_scrollbar.set,
                                               xscrollcommand=h_scrollbar.set)

        # Pack componentes
        self.widgets['records_tree'].pack(side='left', fill='both', expand=True)
        v_scrollbar.pack(side='right', fill='y')
        h_scrollbar.pack(side='bottom', fill='x')

        # Bind evento de selección
        self.widgets['records_tree'].bind('<<TreeviewSelect>>', self._on_record_select)

    def _create_detail_panel(self, parent):
        """Crea el panel de detalle del registro seleccionado"""
        detail_frame = tk.Frame(parent, bg=self.colors['bg_primary'], height=120)
        detail_frame.pack(fill='x')
        detail_frame.pack_propagate(False)

        # Header del panel
        detail_header = tk.Frame(detail_frame, bg=self.colors['bg_secondary'], height=35)
        detail_header.pack(fill='x')
        detail_header.pack_propagate(False)

        tk.Label(detail_header, text="🔍 Detalle del Registro Seleccionado",
                 bg=self.colors['bg_secondary'], fg=self.colors['text_primary'],
                 font=('Arial', 10, 'bold')).pack(side='left', padx=15, pady=8)

        # Contenido del detalle
        detail_content = tk.Frame(detail_frame, bg=self.colors['bg_tertiary'])
        detail_content.pack(fill='both', expand=True, padx=1, pady=(0, 1))

        self.widgets['detail_text'] = scrolledtext.ScrolledText(
            detail_content, height=4, bg=self.colors['bg_tertiary'],
            fg=self.colors['text_primary'], font=('Arial', 9),
            relief='flat', wrap=tk.WORD, state=tk.DISABLED
        )
        self.widgets['detail_text'].pack(fill='both', expand=True, padx=10, pady=8)

        # Mensaje inicial
        self._update_detail_panel("Seleccione un registro para ver los detalles...")

    def _create_styled_entry(self, parent, **kwargs):
        """Crea un Entry con estilo"""
        entry = tk.Entry(
            parent,
            bg=self.colors['bg_tertiary'],
            fg=self.colors['text_primary'],
            font=('Arial', 9),
            relief='flat',
            bd=5,
            **kwargs
        )
        return entry

    def _create_styled_button(self, parent, text, command, color):
        """Crea un botón con estilo"""
        btn = tk.Button(
            parent,
            text=text,
            command=command,
            bg=color,
            fg='white',
            font=('Arial', 9, 'bold'),
            relief='flat',
            padx=15,
            pady=8,
            cursor='hand2'
        )
        return btn

    # ===== MÉTODOS DE REPORTES Y EMAIL =====
    def _get_records_for_report(self, report_type):
        """Obtiene registros según el tipo de reporte seleccionado"""
        if report_type == "Registros Actuales (Filtrados)":
            # Usar filtros actuales
            date_from = self.widgets['date_from'].get().strip()
            date_to = self.widgets['date_to'].get().strip()
            status_filter = self.widgets['status_filter'].get()
            user_filter = self.widgets['user_filter'].get()
            profile_filter = self.widgets['profile_filter'].get()

            return self.registry_manager.get_filtered_records(
                date_from=date_from if date_from else None,
                date_to=date_to if date_to else None,
                status_filter=status_filter,
                user_filter=user_filter,
                profile_filter=profile_filter
            )

        elif report_type == "Últimos 7 días":
            cutoff_date = (datetime.now() - timedelta(days=7)).date().isoformat()
            return self.registry_manager.get_filtered_records(date_from=cutoff_date)

        elif report_type == "Últimos 30 días":
            cutoff_date = (datetime.now() - timedelta(days=30)).date().isoformat()
            return self.registry_manager.get_filtered_records(date_from=cutoff_date)

        elif report_type == "Solo Exitosos":
            return self.registry_manager.get_filtered_records(status_filter="Exitoso")

        elif report_type == "Solo Fallidos":
            return self.registry_manager.get_filtered_records(status_filter="Fallido")

        elif report_type == "Solo Ejecuciones Manuales":
            return self.registry_manager.get_filtered_records(user_filter="Usuario")

        elif report_type == "Solo Ejecuciones Automáticas":
            return self.registry_manager.get_filtered_records(user_filter="Sistema")

        else:  # "Todos los Registros"
            return self.registry_manager.get_all_records()

    def _export_to_excel(self):
        """Exporta registros a Excel"""
        if not EXCEL_AVAILABLE:
            messagebox.showerror("Error", "openpyxl no está instalado.\nInstale con: pip install openpyxl")
            return

        try:
            # Obtener registros según tipo seleccionado
            report_type = self.widgets['report_type'].get()
            records = self._get_records_for_report(report_type)

            if not records:
                messagebox.showwarning("Sin Datos", "No hay registros para exportar según los filtros seleccionados.")
                return

            # Solicitar ubicación de archivo
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"reporte_ejecuciones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )

            if not filename:
                return

            # Generar reporte
            report_title = f"Reporte de Ejecuciones - {report_type}"
            filename, success, message = self.report_generator.export_to_excel(
                records, filename, report_title
            )

            if success:
                messagebox.showinfo("Éxito", f"Reporte exportado exitosamente.\n\n{message}")
            else:
                messagebox.showerror("Error", f"Error exportando reporte:\n{message}")

        except Exception as e:
            messagebox.showerror("Error", f"Error inesperado al exportar:\n{str(e)}")

    def _send_report_email(self):
        """Genera reporte y lo envía por email"""
        if not EXCEL_AVAILABLE:
            messagebox.showerror("Error", "openpyxl no está instalado.\nInstale con: pip install openpyxl")
            return

        if not self.email_tab or not self.email_tab.is_email_configured():
            messagebox.showerror("Email No Configurado",
                                 "Debe configurar el email en la pestaña 'Email' antes de enviar reportes.")
            return

        def send_thread():
            try:
                # Obtener registros
                report_type = self.widgets['report_type'].get()
                records = self._get_records_for_report(report_type)

                if not records:
                    self.frame.after(0, lambda: messagebox.showwarning(
                        "Sin Datos", "No hay registros para enviar según los filtros seleccionados."))
                    return

                # Generar Excel temporal
                temp_filename = f"temp_reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                report_title = f"Reporte de Ejecuciones - {report_type}"

                filename, success, message = self.report_generator.export_to_excel(
                    records, temp_filename, report_title
                )

                if not success:
                    self.frame.after(0, lambda: messagebox.showerror("Error", f"Error generando Excel:\n{message}"))
                    return

                # Preparar email simplificado
                subject = f"Syncro Bot - {report_title} - {datetime.now().strftime('%d/%m/%Y')}"

                stats = {
                    'total': len(records),
                    'exitosos': len([r for r in records if r['estado'] == 'Exitoso']),
                    'fallidos': len([r for r in records if r['estado'] == 'Fallido']),
                    'usuario': len([r for r in records if r['usuario'] == 'Usuario']),
                    'sistema': len([r for r in records if r['usuario'] == 'Sistema'])
                }

                # ===== NUEVO FORMATO SIMPLIFICADO =====
                body = f"""Estimado/a,

Se adjunta el reporte de ejecuciones del sistema Syncro Bot correspondiente a: {report_type}

RESUMEN EJECUTIVO:
Total de Ejecuciones: {stats['total']}
Ejecuciones Exitosas: {stats['exitosos']}
Ejecuciones Fallidas: {stats['fallidos']}
Ejecuciones Manuales: {stats['usuario']}
Ejecuciones Automáticas: {stats['sistema']}
Tasa de Éxito: {(stats['exitosos'] / stats['total'] * 100 if stats['total'] > 0 else 0):.1f}%

Saludos cordiales,
Sistema Syncro Bot"""
                # ===== FIN NUEVO FORMATO =====

                # Enviar email con archivo adjunto
                success_email, message_email = self._send_email_with_attachment(subject, body, filename)

                # Limpiar archivo temporal
                try:
                    if os.path.exists(temp_filename):
                        os.remove(temp_filename)
                except:
                    pass

                # Mostrar resultado
                if success_email:
                    self.frame.after(0, lambda: messagebox.showinfo(
                        "Éxito", f"Reporte enviado exitosamente por email.\n\n{message_email}"))
                else:
                    self.frame.after(0, lambda: messagebox.showerror(
                        "Error", f"Error enviando email:\n{message_email}"))

            except Exception as e:
                # Limpiar archivos temporales en caso de error
                try:
                    if 'temp_filename' in locals() and os.path.exists(temp_filename):
                        os.remove(temp_filename)
                except:
                    pass

                self.frame.after(0, lambda: messagebox.showerror(
                    "Error", f"Error inesperado enviando reporte:\n{str(e)}"))

        # Mostrar mensaje de progreso
        messagebox.showinfo("Enviando", "Generando y enviando reporte por email...\nEsto puede tardar unos momentos.")

        # Ejecutar en hilo separado
        threading.Thread(target=send_thread, daemon=True).start()

    def _send_email_with_attachment(self, subject, body, attachment_path):
        """Envía email con archivo adjunto"""
        try:
            # Verificar que el archivo existe
            if not os.path.exists(attachment_path):
                return False, f"El archivo {attachment_path} no existe"

            # Enviar email con adjunto usando argumentos nombrados
            return self.email_tab.send_email(
                subject=subject,
                body=body,
                attachments=[attachment_path]
            )
        except Exception as e:
            return False, str(e)

    def _update_email_status(self):
        """Actualiza el estado del email en la interfaz"""
        if not self.email_tab:
            self.widgets['email_status'].configure(text="No disponible", fg=self.colors['error'])
        elif self.email_tab.is_email_configured():
            self.widgets['email_status'].configure(text="✅ Configurado", fg=self.colors['success'])
        else:
            self.widgets['email_status'].configure(text="❌ Sin configurar", fg=self.colors['error'])

    # ===== FIN MÉTODOS DE REPORTES =====

    def _apply_filters(self):
        """Aplica los filtros seleccionados"""
        # Obtener valores de filtros
        date_from = self.widgets['date_from'].get().strip()
        date_to = self.widgets['date_to'].get().strip()
        status_filter = self.widgets['status_filter'].get()
        user_filter = self.widgets['user_filter'].get()
        profile_filter = self.widgets['profile_filter'].get()

        # Validar fechas si se proporcionan
        if date_from and not self._validate_date(date_from):
            messagebox.showerror("Error", "Formato de fecha 'Desde' inválido. Use YYYY-MM-DD")
            return

        if date_to and not self._validate_date(date_to):
            messagebox.showerror("Error", "Formato de fecha 'Hasta' inválido. Use YYYY-MM-DD")
            return

        # Aplicar filtros
        try:
            filtered_records = self.registry_manager.get_filtered_records(
                date_from=date_from if date_from else None,
                date_to=date_to if date_to else None,
                status_filter=status_filter,
                user_filter=user_filter,
                profile_filter=profile_filter
            )

            self._populate_records_table(filtered_records)
            messagebox.showinfo("Filtros Aplicados", f"Se encontraron {len(filtered_records)} registros")

        except Exception as e:
            messagebox.showerror("Error", f"Error aplicando filtros:\n{str(e)}")

    def _clear_filters(self):
        """Limpia todos los filtros"""
        self.widgets['date_from'].delete(0, 'end')
        self.widgets['date_to'].delete(0, 'end')
        self.widgets['status_filter'].set("Todos")
        self.widgets['user_filter'].set("Todos")
        self.widgets['profile_filter'].set("Todos")

        # Recargar todos los registros
        self.load_records()

    def _clean_old_records(self):
        """Limpia registros antiguos"""
        if messagebox.askyesno("Confirmar",
                               "¿Está seguro de eliminar registros con más de 30 días?\n\n" +
                               "Esta acción no se puede deshacer."):
            try:
                old_count = len(self.registry_manager.registry)
                self.registry_manager.clear_old_records(30)
                new_count = len(self.registry_manager.registry)
                deleted = old_count - new_count

                messagebox.showinfo("Éxito", f"Se eliminaron {deleted} registros antiguos")
                self.load_records()

            except Exception as e:
                messagebox.showerror("Error", f"Error limpiando registros:\n{str(e)}")

    def _clear_all_records(self):
        """Elimina todos los registros"""
        if messagebox.askyesno("Confirmar Eliminación",
                               "¿Está COMPLETAMENTE SEGURO de eliminar TODOS los registros?\n\n" +
                               "Esta acción eliminará permanentemente todo el historial.\n" +
                               "NO se puede deshacer."):
            try:
                self.registry_manager.clear_registry()
                messagebox.showinfo("Éxito", "Todos los registros han sido eliminados")
                self.load_records()

            except Exception as e:
                messagebox.showerror("Error", f"Error eliminando registros:\n{str(e)}")

    def _validate_date(self, date_str):
        """Valida formato de fecha YYYY-MM-DD"""
        try:
            datetime.strptime(date_str, "%Y-%m-%d")
            return True
        except ValueError:
            return False

    def _on_record_select(self, event):
        """Maneja la selección de un registro en la tabla"""
        selection = self.widgets['records_tree'].selection()
        if not selection:
            self.selected_record = None
            self._update_detail_panel("Seleccione un registro para ver los detalles...")
            return

        # Obtener el registro seleccionado
        item = self.widgets['records_tree'].item(selection[0])
        values = item['values']

        if values:
            # Buscar el registro completo por fecha y hora
            fecha = values[0]
            hora_inicio = values[1]

            records = self.registry_manager.get_all_records()
            for record in records:
                if record['fecha'] == fecha and record['hora_inicio'] == hora_inicio:
                    self.selected_record = record
                    self._show_record_detail(record)
                    break

    def _show_record_detail(self, record):
        """Muestra el detalle completo de un registro"""
        detail_text = f"""ID: {record['id']}
Fecha: {record['fecha']}
Hora Inicio: {record['hora_inicio']}
Hora Fin: {record['hora_fin']}
Perfil: {record['perfil']}
Duración: {record['duracion']}
Estado: {record['estado']}
Usuario: {record['usuario']}"""

        if record['error_message']:
            detail_text += f"\nError: {record['error_message']}"

        detail_text += f"\nCreado: {record['created'][:19]}"

        self._update_detail_panel(detail_text)

    def _update_detail_panel(self, text):
        """Actualiza el panel de detalle"""
        self.widgets['detail_text'].configure(state=tk.NORMAL)
        self.widgets['detail_text'].delete(1.0, tk.END)
        self.widgets['detail_text'].insert(tk.END, text)
        self.widgets['detail_text'].configure(state=tk.DISABLED)

    def _populate_records_table(self, records):
        """Puebla la tabla con registros"""
        # Limpiar tabla actual
        for item in self.widgets['records_tree'].get_children():
            self.widgets['records_tree'].delete(item)

        # Insertar registros
        for record in records:
            # Formatear estado con emoji
            estado = record['estado']
            if estado == "Exitoso":
                estado = "✅ Exitoso"
            elif estado == "Fallido":
                estado = "❌ Fallido"
            elif estado == "En Ejecución":
                estado = "⏳ En Progreso"

            # Formatear usuario
            usuario = record['usuario']
            if usuario == "Sistema":
                usuario = "🤖 Sistema"
            else:
                usuario = "👤 Usuario"

            self.widgets['records_tree'].insert('', 'end', values=(
                record['fecha'],
                record['hora_inicio'],
                record['hora_fin'],
                record['perfil'],
                record['duracion'],
                estado,
                usuario
            ))

        # Actualizar contador
        self.widgets['record_count'].configure(text=f"{len(records)} registros")

    def _update_statistics(self):
        """Actualiza las estadísticas mostradas"""
        stats = self.registry_manager.get_statistics()

        self.widgets['total_stat'].configure(text=str(stats['total']))
        self.widgets['success_stat'].configure(text=str(stats['successful']))
        self.widgets['failed_stat'].configure(text=str(stats['failed']))
        self.widgets['progress_stat'].configure(text=str(stats['in_progress']))
        self.widgets['manual_stat'].configure(text=str(stats['manual']))
        self.widgets['system_stat'].configure(text=str(stats['system']))
        self.widgets['success_rate_stat'].configure(text=f"{stats['success_rate']:.1f}%")

    def _update_profile_filter_options(self):
        """Actualiza las opciones del filtro de perfil"""
        records = self.registry_manager.get_all_records()
        profiles = set(["Todos"])

        for record in records:
            profiles.add(record['perfil'])

        self.widgets['profile_filter']['values'] = sorted(list(profiles))

    def load_records(self):
        """Carga y muestra todos los registros"""
        try:
            # Cargar registros
            records = self.registry_manager.get_all_records()

            # Actualizar tabla
            self._populate_records_table(records)

            # Actualizar estadísticas
            self._update_statistics()

            # Actualizar opciones de filtro de perfil
            self._update_profile_filter_options()

            # Actualizar estado del email
            self._update_email_status()

            # Limpiar detalle
            self._update_detail_panel("Seleccione un registro para ver los detalles...")

        except Exception as e:
            messagebox.showerror("Error", f"Error cargando registros:\n{str(e)}")

    def add_execution_record(self, start_time, profile_name="Manual", user_type="Usuario"):
        """Añade un nuevo registro de ejecución (método público para otras pestañas)"""
        return self.registry_manager.add_execution_record(
            start_time=start_time,
            profile_name=profile_name,
            user_type=user_type
        )

    def update_execution_record(self, record_id, end_time, status, error_message=""):
        """Actualiza un registro de ejecución (método público para otras pestañas)"""
        result = self.registry_manager.update_execution_record(
            record_id, end_time, status, error_message
        )

        # Refrescar la vista si es necesario
        if hasattr(self, 'widgets') and 'records_tree' in self.widgets:
            self.load_records()

        return result

    # ===== MÉTODOS PÚBLICOS PARA REPORTES AUTOMÁTICOS =====
    def generate_and_send_report(self, report_type="Últimos 7 días", custom_title=None):
        """
        Método público para generar y enviar reportes automáticamente.
        Usado por perfiles de automatización.
        """
        if not EXCEL_AVAILABLE or not self.email_tab or not self.email_tab.is_email_configured():
            return False, "Excel o Email no configurados"

        try:
            # Obtener registros
            records = self._get_records_for_report(report_type)

            if not records:
                return False, "No hay registros para el reporte"

            # Generar Excel temporal
            temp_filename = f"temp_reporte_auto_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            report_title = custom_title or f"Reporte Automático - {report_type}"

            filename, success, message = self.report_generator.export_to_excel(
                records, temp_filename, report_title
            )

            if not success:
                return False, f"Error generando Excel: {message}"

            # Preparar email simplificado
            subject = f"Syncro Bot - {report_title} - {datetime.now().strftime('%d/%m/%Y')}"

            stats = {
                'total': len(records),
                'exitosos': len([r for r in records if r['estado'] == 'Exitoso']),
                'fallidos': len([r for r in records if r['estado'] == 'Fallido']),
                'usuario': len([r for r in records if r['usuario'] == 'Usuario']),
                'sistema': len([r for r in records if r['usuario'] == 'Sistema'])
            }

            # ===== FORMATO SIMPLIFICADO PARA REPORTES AUTOMÁTICOS =====
            body = f"""Estimado/a,

Se adjunta el reporte de ejecuciones del sistema Syncro Bot correspondiente a: {report_type}

RESUMEN EJECUTIVO:
Total de Ejecuciones: {stats['total']}
Ejecuciones Exitosas: {stats['exitosos']}
Ejecuciones Fallidas: {stats['fallidos']}
Ejecuciones Manuales: {stats['usuario']}
Ejecuciones Automáticas: {stats['sistema']}
Tasa de Éxito: {(stats['exitosos'] / stats['total'] * 100 if stats['total'] > 0 else 0):.1f}%

Saludos cordiales,
Sistema Syncro Bot"""
            # ===== FIN FORMATO SIMPLIFICADO =====

            # Enviar email
            success_email, message_email = self._send_email_with_attachment(subject, body, filename)

            # Limpiar archivo temporal
            try:
                if os.path.exists(temp_filename):
                    os.remove(temp_filename)
            except:
                pass

            return success_email, message_email

        except Exception as e:
            return False, str(e)

    # ===== FIN MÉTODOS PÚBLICOS =====

    def cleanup(self):
        """Limpia recursos al cerrar"""
        try:
            self.report_generator.cleanup_temp_files()
        except:
            pass