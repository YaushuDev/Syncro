# excel_exporter.py
# Ubicación: /syncro_bot/gui/components/automation/handlers/excel_exporter.py
"""
Exportador especializado de datos a archivos Excel con soporte para números
de serie de equipos y fecha de creación. Maneja la creación de archivos Excel con
formato profesional incluyendo todas las columnas extraídas.
"""

import os
from datetime import datetime
from typing import List, Dict, Optional

# Importaciones para Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl no está instalado. Funcionalidad de Excel deshabilitada.")
    print("Instale con: pip install openpyxl")


class ExcelExporter:
    """Exportador especializado de datos a archivos Excel con soporte para números de serie y fecha creación"""

    def __init__(self, logger=None):
        self.logger = logger

        # Configuración de archivos
        self.default_filename = "datos_extraidos_syncro_bot"
        self.output_directory = "reportes_excel"

        # Mapeo de nombres de columnas actualizado con número de serie
        self.column_headers = {
            'fila_numero': 'Fila #',
            'numero_orden': 'Número de Orden',
            'cliente': 'Cliente',
            'numero_serie': 'Número de Serie',  # Cambiado de telefono_cliente
            'tecnico': 'Técnico',
            'distrito': 'Distrito',
            'barrio': 'Barrio',
            'canton': 'Cantón',
            'fecha_creacion': 'Fecha Creación',
            'observaciones': 'Observaciones',
            'estado': 'Estado',
            'despacho': 'Despacho'
        }

        # Orden preferido de columnas actualizado con número de serie
        self.column_order = [
            'fila_numero', 'numero_orden', 'cliente', 'numero_serie',  # numero_serie en lugar de telefono_cliente
            'tecnico', 'distrito', 'barrio', 'canton', 'fecha_creacion',
            'estado', 'despacho', 'observaciones'
        ]

    def _log(self, message, level="INFO"):
        """Log interno con fallback"""
        if self.logger:
            self.logger(message, level)
        else:
            print(f"[{level}] {message}")

    def is_available(self) -> bool:
        """Verifica si el exportador está disponible"""
        return OPENPYXL_AVAILABLE

    def export_to_excel(self, data: List[Dict], filename: Optional[str] = None) -> tuple[bool, str, str]:
        """Exporta datos a archivo Excel con formato profesional incluyendo números de serie y fecha creación"""
        try:
            if not OPENPYXL_AVAILABLE:
                return False, "openpyxl no está disponible", ""

            if not data:
                return False, "No hay datos para exportar", ""

            # 🔍 DEBUG: Verificar datos recibidos
            self._log(f"🔍 DEBUG: Recibidos {len(data)} registros para exportar")

            # Contar registros con número de serie y fecha
            series_count = 0
            dates_count = 0
            for record in data:
                numero_serie = record.get('numero_serie', '')
                if numero_serie and numero_serie not in ['Sin número de serie', 'Error extracción', 'Error popup', 'Campo no encontrado']:
                    series_count += 1

                fecha = record.get('fecha_creacion', '')
                if fecha and fecha.strip():
                    dates_count += 1

            self._log(f"🔢 {series_count} registros con número de serie de {len(data)} totales")
            self._log(f"📅 {dates_count} registros with fecha creación de {len(data)} totales")

            # Preparar directorio de salida
            if not os.path.exists(self.output_directory):
                os.makedirs(self.output_directory)
                self._log(f"📁 Directorio creado: {self.output_directory}")

            # Generar nombre de archivo
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{self.default_filename}_{timestamp}.xlsx"

            if not filename.endswith('.xlsx'):
                filename += '.xlsx'

            filepath = os.path.join(self.output_directory, filename)

            self._log(f"📊 Iniciando exportación a Excel uniforme: {filepath}")

            # Crear workbook y worksheet
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Datos Extraídos"

            # Configurar datos para exportación
            success = self._setup_worksheet(worksheet, data)
            if not success:
                return False, "Error configurando worksheet", ""

            # Guardar archivo
            workbook.save(filepath)
            self._log(f"✅ Archivo Excel guardado exitosamente: {filepath}")

            # Verificar que el archivo se creó correctamente
            if os.path.exists(filepath):
                file_size = os.path.getsize(filepath)
                success_message = f"Excel creado: {len(data)} registros exportados ({series_count} con número de serie, {dates_count} con fecha, {file_size} bytes)"
                return True, success_message, filepath
            else:
                return False, "Archivo no se creó correctamente", ""

        except Exception as e:
            error_msg = f"Error creando archivo Excel: {str(e)}"
            self._log(error_msg, "ERROR")
            return False, error_msg, ""

    def _setup_worksheet(self, worksheet, data: List[Dict]) -> bool:
        """Configura el worksheet con datos y formato incluyendo números de serie y fecha creación"""
        try:
            # 🔍 DEBUG: Verificar datos en setup
            self._log(f"🔍 DEBUG: Configurando worksheet con {len(data)} registros")

            # Determinar columnas a incluir basándose en los datos
            columns_to_include = self._determine_columns_improved(data)
            self._log(f"📋 Columnas incluidas: {columns_to_include}")

            if not columns_to_include:
                self._log("❌ ERROR: No se determinaron columnas para incluir", "ERROR")
                return False

            # Crear headers
            self._create_headers(worksheet, columns_to_include)

            # Insertar datos
            rows_inserted = self._insert_data_improved(worksheet, data, columns_to_include)
            self._log(f"📊 {rows_inserted} filas de datos insertadas en Excel")

            if rows_inserted == 0:
                self._log("⚠️ WARNING: No se insertaron datos en el Excel", "WARNING")

            # Aplicar formato
            self._apply_formatting(worksheet, len(data), len(columns_to_include))

            # Crear tabla solo si hay datos
            if rows_inserted > 0:
                self._create_table(worksheet, rows_inserted, columns_to_include)

            # Ajustar ancho de columnas (incluyendo número de serie y fecha)
            self._adjust_column_widths(worksheet, columns_to_include)

            return True

        except Exception as e:
            self._log(f"Error configurando worksheet: {str(e)}", "ERROR")
            return False

    def _determine_columns_improved(self, data: List[Dict]) -> List[str]:
        """🔧 Determina qué columnas incluir dando prioridad al número de serie y fecha creación"""
        if not data:
            self._log("⚠️ No hay datos para determinar columnas", "WARNING")
            return []

        # Obtener todas las columnas presentes en los datos
        all_columns = set()
        for record in data:
            all_columns.update(record.keys())

        self._log(f"🔍 DEBUG: Columnas encontradas en datos: {list(all_columns)}")

        # SIEMPRE incluir estas columnas principales si existen
        always_include = ['fila_numero', 'numero_orden', 'cliente']
        columns_with_data = []

        # Primero agregar las columnas que siempre queremos
        for col in always_include:
            if col in all_columns:
                columns_with_data.append(col)
                self._log(f"✅ Columna obligatoria incluida: {col}")

        # 🆕 VERIFICAR NÚMERO DE SERIE ESPECÍFICAMENTE
        if 'numero_serie' in all_columns:
            # Verificar si hay números de serie válidos
            valid_series = 0
            for record in data:
                numero_serie = record.get('numero_serie', '')
                if numero_serie and numero_serie not in ['Sin número de serie', 'Error extracción', 'Error popup',
                                                        'Campo no encontrado', 'Error lectura popup', 'Sin tabla popup', '']:
                    valid_series += 1

            if valid_series > 0:
                columns_with_data.append('numero_serie')
                self._log(f"🔢 ✅ Columna número de serie incluida: {valid_series} números de serie válidos")
            else:
                self._log(f"🔢 ⚠️ Columna número de serie omitida: sin números de serie válidos")

        # VERIFICAR FECHA CREACIÓN ESPECÍFICAMENTE
        if 'fecha_creacion' in all_columns:
            # Verificar si hay fechas válidas
            valid_dates = 0
            for record in data:
                fecha = record.get('fecha_creacion', '')
                if fecha and str(fecha).strip() and str(fecha).strip() != '':
                    valid_dates += 1

            if valid_dates > 0:
                columns_with_data.append('fecha_creacion')
                self._log(f"📅 ✅ Columna fecha creación incluida: {valid_dates} fechas válidas")
            else:
                self._log(f"📅 ⚠️ Columna fecha creación omitida: sin fechas válidas")

        # Luego verificar otras columnas según el orden preferido
        for col in self.column_order:
            if col in all_columns and col not in columns_with_data:
                # Verificar si la columna tiene al menos algunos datos NO VACÍOS
                has_meaningful_data = False

                for record in data:
                    value = record.get(col, '')
                    # Convertir a string y limpiar para verificar
                    str_value = str(value).strip() if value is not None else ''
                    if str_value and str_value != 'None' and str_value != '':
                        has_meaningful_data = True
                        break

                if has_meaningful_data:
                    columns_with_data.append(col)
                    self._log(f"✅ Columna con datos incluida: {col}")
                else:
                    self._log(f"⚠️ Columna sin datos omitida: {col}")

        # Agregar columnas adicionales que no estén en el orden predefinido
        for col in all_columns:
            if col not in columns_with_data and col in self.column_headers:
                # Verificar si tiene datos
                has_data = any(
                    str(record.get(col, '')).strip()
                    for record in data
                    if record.get(col) is not None
                )
                if has_data:
                    columns_with_data.append(col)
                    self._log(f"✅ Columna adicional incluida: {col}")

        if not columns_with_data:
            # Si no se detectaron columnas con datos, incluir al menos las básicas
            basic_columns = [col for col in ['fila_numero', 'numero_orden', 'cliente', 'tecnico']
                             if col in all_columns]
            columns_with_data = basic_columns
            self._log(f"⚠️ Fallback: usando columnas básicas: {columns_with_data}")

        self._log(f"📋 Columnas finales a incluir: {columns_with_data}")
        return columns_with_data

    def _create_headers(self, worksheet, columns: List[str]):
        """Crea los headers del Excel con formato uniforme"""
        self._log(f"📋 Creando headers uniformes para {len(columns)} columnas")

        for col_index, column in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col_index)
            header_text = self.column_headers.get(column, column.title())
            cell.value = header_text

            # 🎨 Estilo uniforme para todos los headers
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # Azul uniforme

            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

            self._log(f"📋 Header uniforme {col_index}: '{header_text}' para columna '{column}'")

        self._log("📋 Headers uniformes creados exitosamente")

    def _insert_data_improved(self, worksheet, data: List[Dict], columns: List[str]) -> int:
        """🔧 Inserta los datos en el worksheet con formato uniforme"""
        self._log(f"📊 Insertando datos: {len(data)} registros, {len(columns)} columnas")

        rows_inserted = 0

        for row_index, record in enumerate(data, 2):  # Empezar en fila 2
            row_has_data = False

            for col_index, column in enumerate(columns, 1):
                cell = worksheet.cell(row=row_index, column=col_index)

                # Obtener valor del registro
                raw_value = record.get(column, '')

                # Procesamiento específico por tipo de columna
                if column == 'numero_serie':
                    processed_value = self._process_serie_value(raw_value)
                elif column == 'fecha_creacion':
                    processed_value = self._process_date_value(raw_value)
                else:
                    processed_value = self._process_cell_value(raw_value)

                # Asignar valor a la celda
                cell.value = processed_value

                # Verificar si esta fila tiene al menos un dato significativo
                if processed_value and str(processed_value).strip():
                    row_has_data = True

                # 🔍 DEBUG: Log para las primeras filas
                if row_index <= 3:
                    if column == 'numero_serie':
                        self._log(f"🔢 DEBUG: Fila {row_index}, Número Serie: '{raw_value}' → '{processed_value}'")
                    elif column == 'fecha_creacion':
                        self._log(f"📅 DEBUG: Fila {row_index}, Fecha: '{raw_value}' → '{processed_value}'")
                    elif column in ['numero_orden', 'cliente']:
                        self._log(f"🔍 DEBUG: Fila {row_index}, {column}: '{raw_value}' → '{processed_value}'")

            if row_has_data:
                rows_inserted += 1
            else:
                self._log(f"⚠️ Fila {row_index} sin datos significativos")

        self._log(f"📊 Proceso completado: {rows_inserted} filas con datos de {len(data)} totales")
        return rows_inserted

    def _process_serie_value(self, raw_value):
        """🆕 Procesa específicamente valores de número de serie"""
        if not raw_value:
            return "Sin número de serie"

        str_value = str(raw_value).strip()

        # Casos de error específicos de número de serie
        error_cases = [
            'Sin número de serie', 'Error extracción', 'Error popup', 'Error lectura popup',
            'Sin tabla popup', 'Campo no encontrado', 'none', 'null'
        ]

        if str_value.lower() in [case.lower() for case in error_cases]:
            return "Error"

        # Limpiar el número de serie
        cleaned = str_value.replace('&nbsp;', ' ')
        cleaned = cleaned.replace('\xa0', ' ')
        cleaned = cleaned.replace('\u00a0', ' ')
        cleaned = ' '.join(cleaned.split())

        return cleaned if cleaned else "Sin número de serie"

    def _process_date_value(self, raw_value):
        """Procesa específicamente valores de fecha creación"""
        if not raw_value:
            return "Sin fecha"

        str_value = str(raw_value).strip()

        # Casos de valores vacíos
        if str_value.lower() in ['none', 'null', '', '&nbsp;']:
            return "Sin fecha"

        # Limpiar la fecha
        cleaned = str_value.replace('&nbsp;', ' ')
        cleaned = cleaned.replace('\xa0', ' ')
        cleaned = cleaned.replace('\u00a0', ' ')
        cleaned = ' '.join(cleaned.split())

        return cleaned if cleaned else "Sin fecha"

    def _process_cell_value(self, raw_value):
        """🔧 Procesa y limpia valores de celda de manera robusta"""
        # Manejar None
        if raw_value is None:
            return ""

        # Convertir a string
        str_value = str(raw_value)

        # Limpiar espacios y caracteres especiales
        cleaned_value = str_value.strip()

        # Limpiar valores que indican vacío
        if cleaned_value.lower() in ['none', 'null', '&nbsp;', '']:
            return ""

        # Limpiar caracteres especiales HTML
        cleaned_value = cleaned_value.replace('&nbsp;', ' ')
        cleaned_value = cleaned_value.replace('\xa0', ' ')
        cleaned_value = cleaned_value.replace('\u00a0', ' ')

        # Limpiar espacios múltiples
        cleaned_value = ' '.join(cleaned_value.split())

        return cleaned_value if cleaned_value else ""

    def _apply_formatting(self, worksheet, num_rows: int, num_cols: int):
        """Aplica formato general al worksheet"""
        # Borders para todas las celdas con datos
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Aplicar bordes y alineación a todas las celdas con datos
        for row in range(1, num_rows + 2):  # +2 porque empezamos en 1 y agregamos header
            for col in range(1, num_cols + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border

                # Alineación para celdas de datos (no headers)
                if row > 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        self._log("🎨 Formato aplicado exitosamente")

    def _create_table(self, worksheet, num_rows: int, columns: List[str]):
        """Crea una tabla de Excel para mejor visualización"""
        try:
            if num_rows == 0:
                self._log("⚠️ No se puede crear tabla: sin filas de datos")
                return

            # Definir rango de la tabla
            start_cell = "A1"
            end_cell = f"{get_column_letter(len(columns))}{num_rows + 1}"
            table_range = f"{start_cell}:{end_cell}"

            self._log(f"📋 Creando tabla Excel en rango: {table_range}")

            # Crear tabla
            table = Table(displayName="DatosExtraidos", ref=table_range)

            # Estilo de tabla
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style

            # Agregar tabla al worksheet
            worksheet.add_table(table)
            self._log("📋 Tabla de Excel creada exitosamente")

        except Exception as e:
            self._log(f"⚠️ No se pudo crear tabla: {str(e)}", "WARNING")

    def _adjust_column_widths(self, worksheet, columns: List[str]):
        """Ajusta automáticamente el ancho de las columnas con formato uniforme"""
        try:
            # Definir anchos mínimos y máximos
            min_width = 10
            max_width = 50

            for col_index, column in enumerate(columns, 1):
                column_letter = get_column_letter(col_index)

                # Ancho basado en el tipo de columna
                if column == 'fila_numero':
                    width = 8
                elif column == 'numero_orden':
                    width = 18
                elif column == 'cliente':
                    width = 25
                elif column == 'numero_serie':  # Cambiado de telefono_cliente
                    width = 20
                elif column == 'fecha_creacion':
                    width = 18
                elif column in ['tecnico', 'distrito', 'barrio', 'canton']:
                    width = 15
                elif column == 'observaciones':
                    width = 30
                else:
                    width = 12

                # Aplicar límites
                width = max(min_width, min(width, max_width))
                worksheet.column_dimensions[column_letter].width = width

            self._log("📏 Anchos de columna ajustados con formato uniforme")

        except Exception as e:
            self._log(f"⚠️ Error ajustando anchos: {str(e)}", "WARNING")

    def create_summary_sheet(self, workbook, data: List[Dict], summary_info: Dict):
        """Crea una hoja de resumen con estadísticas y formato uniforme"""
        try:
            summary_sheet = workbook.create_sheet("Resumen")

            # Título
            summary_sheet['A1'] = "Resumen de Extracción de Datos"
            summary_sheet['A1'].font = Font(size=16, bold=True)

            # Información general
            row = 3
            info_items = [
                ("Fecha de extracción:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                ("Total de registros:", len(data)),
                ("Registros válidos:", summary_info.get('valid_records', 0)),
                ("Números de serie extraídos:", summary_info.get('series_extracted', 0)),  # Cambiado
                ("Fechas creación extraídas:", self._count_valid_dates(data)),
                ("Errores de número de serie:", summary_info.get('series_errors', 0)),  # Cambiado
                ("Método de extracción:", summary_info.get('extraction_method', 'N/A')),  # Cambiado
                ("Campos extraídos:", ', '.join(summary_info.get('fields_extracted', []))),
            ]

            for label, value in info_items:
                summary_sheet[f'A{row}'] = label
                summary_sheet[f'A{row}'].font = Font(bold=True)

                # 🎨 Sin colores especiales - formato uniforme
                summary_sheet[f'B{row}'] = value
                row += 1

            # Estadísticas de tasa de éxito
            if summary_info.get('series_extracted', 0) > 0:  # Cambiado
                row += 1
                summary_sheet[f'A{row}'] = "Tasa de éxito de números de serie:"  # Cambiado
                summary_sheet[f'A{row}'].font = Font(bold=True)

                total_attempts = summary_info.get('series_extracted', 0) + summary_info.get('series_errors', 0)  # Cambiado
                success_rate = (
                        summary_info.get('series_extracted', 0) / total_attempts * 100) if total_attempts > 0 else 0  # Cambiado
                summary_sheet[f'B{row}'] = f"{success_rate:.1f}%"
                row += 1

            # Estadísticas por técnico si están disponibles
            if 'tecnicos_count' in summary_info:
                row += 2
                summary_sheet[f'A{row}'] = "Distribución por Técnico:"
                summary_sheet[f'A{row}'].font = Font(bold=True)
                row += 1

                for tecnico, count in summary_info['tecnicos_count'].items():
                    summary_sheet[f'A{row}'] = tecnico
                    summary_sheet[f'B{row}'] = count
                    row += 1

            # Ajustar anchos
            summary_sheet.column_dimensions['A'].width = 25
            summary_sheet.column_dimensions['B'].width = 30

            self._log("📊 Hoja de resumen con formato uniforme creada")

        except Exception as e:
            self._log(f"⚠️ Error creando resumen: {str(e)}", "WARNING")

    def _count_valid_dates(self, data: List[Dict]) -> int:
        """Cuenta las fechas de creación válidas"""
        try:
            valid_dates = 0
            for record in data:
                fecha = record.get('fecha_creacion', '')
                if fecha and str(fecha).strip() and str(fecha).strip() not in ['Sin fecha', '']:
                    valid_dates += 1
            return valid_dates
        except Exception:
            return 0

    def export_with_summary(self, data: List[Dict], summary_info: Dict,
                            filename: Optional[str] = None) -> tuple[bool, str, str]:
        """Exporta datos con hoja de resumen incluida (con estadísticas completas)"""
        try:
            if not OPENPYXL_AVAILABLE:
                return False, "openpyxl no está disponible", ""

            if not data:
                return False, "No hay datos para exportar", ""

            # 🔍 DEBUG: Verificar datos antes de exportar
            series_count = summary_info.get('series_extracted', 0)  # Cambiado
            dates_count = self._count_valid_dates(data)
            self._log(
                f"🔍 DEBUG: Exportando con resumen - {len(data)} registros, {series_count} números de serie, {dates_count} fechas")

            # Preparar archivo
            if not os.path.exists(self.output_directory):
                os.makedirs(self.output_directory)

            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{self.default_filename}_con_resumen_{timestamp}.xlsx"

            if not filename.endswith('.xlsx'):
                filename += '.xlsx'

            filepath = os.path.join(self.output_directory, filename)

            # Crear workbook
            workbook = openpyxl.Workbook()

            # Hoja de datos principal
            main_sheet = workbook.active
            main_sheet.title = "Datos Extraídos"
            setup_success = self._setup_worksheet(main_sheet, data)

            if not setup_success:
                self._log("❌ Error configurando hoja principal", "ERROR")

            # Hoja de resumen
            self.create_summary_sheet(workbook, data, summary_info)

            # Guardar
            workbook.save(filepath)

            if os.path.exists(filepath):
                file_size = os.path.getsize(filepath)
                success_message = f"Excel con resumen creado: {len(data)} registros, {series_count} números de serie, {dates_count} fechas ({file_size} bytes)"
                self._log(f"✅ {success_message}")
                return True, success_message, filepath
            else:
                return False, "Archivo no se creó correctamente", ""

        except Exception as e:
            error_msg = f"Error creando Excel con resumen: {str(e)}"
            self._log(error_msg, "ERROR")
            return False, error_msg, ""

    def validate_excel_file(self, filepath: str) -> tuple[bool, str]:
        """Valida que el archivo Excel se creó correctamente incluyendo números de serie y fechas"""
        try:
            if not os.path.exists(filepath):
                return False, "Archivo no existe"

            # Intentar abrir el archivo
            workbook = openpyxl.load_workbook(filepath)
            worksheet = workbook.active

            # Verificar que tiene datos
            if worksheet.max_row <= 1:
                return False, "Archivo no contiene datos (solo headers)"

            # Verificar que tiene headers
            first_row_values = [cell.value for cell in worksheet[1]]
            if not any(first_row_values):
                return False, "Archivo no tiene headers válidos"

            # 🆕 Verificar columnas especiales
            has_serie_column = any('Número de Serie' in str(val) for val in first_row_values if val)  # Cambiado
            has_date_column = any('Fecha Creación' in str(val) for val in first_row_values if val)

            # Verificar que hay datos en las celdas
            data_cells_found = 0
            serie_cells_found = 0  # Cambiado
            date_cells_found = 0

            for row in range(2, min(worksheet.max_row + 1, 5)):  # Verificar primeras filas
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value and str(cell_value).strip():
                        data_cells_found += 1

                        # 🆕 Contar celdas especiales
                        header_value = worksheet.cell(row=1, column=col).value
                        if header_value:
                            if 'Número de Serie' in str(header_value):  # Cambiado
                                serie_cells_found += 1
                            elif 'Fecha Creación' in str(header_value):
                                date_cells_found += 1

            workbook.close()

            file_size = os.path.getsize(filepath)

            if data_cells_found > 0:
                validation_msg = f"Archivo válido ({file_size} bytes, {worksheet.max_row - 1} registros, {data_cells_found} celdas con datos"
                if has_serie_column:  # Cambiado
                    validation_msg += f", columna número de serie incluida con {serie_cells_found} valores"
                if has_date_column:
                    validation_msg += f", columna fecha creación incluida con {date_cells_found} valores"
                validation_msg += ")"
                return True, validation_msg
            else:
                return False, f"Archivo creado pero sin datos en las celdas ({file_size} bytes)"

        except Exception as e:
            return False, f"Error validando archivo: {str(e)}"

    def get_export_info(self) -> Dict:
        """Obtiene información sobre el exportador incluyendo soporte para números de serie y fechas"""
        return {
            'available': OPENPYXL_AVAILABLE,
            'output_directory': self.output_directory,
            'supported_formats': ['xlsx'],
            'default_filename': self.default_filename,
            'column_headers': self.column_headers,
            'serie_support': True,  # Cambiado de phone_support
            'date_support': True,
            'special_formatting': True
        }