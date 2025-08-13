# registry_reports.py
# Ubicación: /syncro_bot/gui/components/registry_reports.py
"""
Sistema completo de generación y envío de reportes para el registro de ejecuciones.
Maneja la generación de archivos Excel, envío por email, reportes automáticos
y diferentes tipos de reportes con integración completa al sistema de email.
"""

import os
import threading
from datetime import datetime
from tkinter import filedialog

# Importaciones para Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl no está instalado. Funcionalidad Excel deshabilitada.")
    print("Instale con: pip install openpyxl")


class ReportTypes:
    """Definiciones de tipos de reportes disponibles"""
    CURRENT_FILTERED = "Registros Actuales (Filtrados)"
    LAST_7_DAYS = "Últimos 7 días"
    LAST_30_DAYS = "Últimos 30 días"
    ONLY_SUCCESSFUL = "Solo Exitosos"
    ONLY_FAILED = "Solo Fallidos"
    ONLY_MANUAL = "Solo Ejecuciones Manuales"
    ONLY_AUTOMATIC = "Solo Ejecuciones Automáticas"
    ALL_RECORDS = "Todos los Registros"

    @classmethod
    def get_all_types(cls):
        """Obtiene todos los tipos de reportes disponibles"""
        return [
            cls.CURRENT_FILTERED, cls.LAST_7_DAYS, cls.LAST_30_DAYS,
            cls.ONLY_SUCCESSFUL, cls.ONLY_FAILED, cls.ONLY_MANUAL,
            cls.ONLY_AUTOMATIC, cls.ALL_RECORDS
        ]


class ReportGenerator:
    """Generador de reportes en Excel con formato profesional integrado"""

    def __init__(self):
        self.temp_files = []

    def is_excel_available(self):
        """Verifica si Excel está disponible"""
        try:
            import openpyxl
            return True
        except ImportError:
            return False

    def export_to_excel(self, records, filename=None, report_title="Reporte de Ejecuciones"):
        """Exporta registros a Excel con formato profesional"""
        # Verificar openpyxl directamente cada vez
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ImportError("openpyxl no está disponible. Instale con: pip install openpyxl")

        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"reporte_ejecuciones_{timestamp}.xlsx"

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte de Ejecuciones"

            self._apply_excel_styles(ws, records, report_title)
            wb.save(filename)

            if filename.startswith("temp_"):
                self.temp_files.append(filename)

            return filename, True, f"Archivo Excel generado: {filename}"

        except Exception as e:
            return None, False, f"Error generando Excel: {str(e)}"

    def _apply_excel_styles(self, ws, records, report_title):
        """Aplica estilos y formatos al archivo Excel"""
        # Importar estilos directamente para asegurar disponibilidad
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Título del reporte
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = report_title
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = center_alignment

        # Fecha de generación
        ws.merge_cells('A2:H2')
        date_cell = ws['A2']
        date_cell.value = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        date_cell.alignment = center_alignment

        ws['A3'] = ""

        # Headers
        headers = ['Fecha', 'Hora Inicio', 'Hora Fin', 'Perfil', 'Duración', 'Estado', 'Usuario', 'Error']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border

        # Datos
        for row_idx, record in enumerate(records, 5):
            self._add_record_row(ws, row_idx, record, border)

        # Ajustar ancho de columnas
        column_widths = [12, 12, 12, 15, 12, 12, 10, 30]
        for col, width in enumerate(column_widths, 1):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(col)].width = width

        # Agregar estadísticas
        self._add_statistics_section(ws, records, len(records) + 6)

    def _add_record_row(self, ws, row_idx, record, border):
        """Añade una fila de registro con formato"""
        from openpyxl.styles import PatternFill

        ws.cell(row=row_idx, column=1, value=record['fecha']).border = border
        ws.cell(row=row_idx, column=2, value=record['hora_inicio']).border = border
        ws.cell(row=row_idx, column=3, value=record['hora_fin']).border = border
        ws.cell(row=row_idx, column=4, value=record['perfil']).border = border
        ws.cell(row=row_idx, column=5, value=record['duracion']).border = border

        # Estado con color
        estado_cell = ws.cell(row=row_idx, column=6, value=record['estado'])
        estado_cell.border = border
        if record['estado'] == 'Exitoso':
            estado_cell.fill = PatternFill(start_color="D5EDDA", end_color="D5EDDA", fill_type="solid")
        elif record['estado'] == 'Fallido':
            estado_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        elif record['estado'] == 'En Ejecución':
            estado_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

        ws.cell(row=row_idx, column=7, value=record['usuario']).border = border
        ws.cell(row=row_idx, column=8, value=record.get('error_message', '')).border = border

    def _add_statistics_section(self, ws, records, start_row):
        """Añade sección de estadísticas al Excel"""
        if not records:
            return

        # Importar estilos directamente
        from openpyxl.styles import Font, PatternFill, Alignment

        total = len(records)
        successful = len([r for r in records if r['estado'] == 'Exitoso'])
        failed = len([r for r in records if r['estado'] == 'Fallido'])
        in_progress = len([r for r in records if r['estado'] == 'En Ejecución'])
        manual = len([r for r in records if r['usuario'] == 'Usuario'])
        system = len([r for r in records if r['usuario'] == 'Sistema'])
        success_rate = (successful / total * 100) if total > 0 else 0

        stats_font = Font(bold=True)
        stats_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

        ws.merge_cells(f'A{start_row}:H{start_row}')
        stats_title = ws[f'A{start_row}']
        stats_title.value = "ESTADÍSTICAS DEL REPORTE"
        stats_title.font = Font(bold=True, size=14)
        stats_title.alignment = Alignment(horizontal="center")
        stats_title.fill = stats_fill

        stats_data = [
            ("Total de Ejecuciones:", total),
            ("Ejecuciones Exitosas:", successful),
            ("Ejecuciones Fallidas:", failed),
            ("En Progreso:", in_progress),
            ("Ejecuciones Manuales:", manual),
            ("Ejecuciones Automáticas:", system),
            ("Tasa de Éxito:", f"{success_rate:.1f}%")
        ]

        for i, (label, value) in enumerate(stats_data, 1):
            label_cell = ws.cell(row=start_row + i, column=1, value=label)
            value_cell = ws.cell(row=start_row + i, column=2, value=value)
            label_cell.font = stats_font
            value_cell.font = Font(bold=True, color="366092")

    def create_temp_filename(self, prefix="temp_reporte"):
        """Crea nombre de archivo temporal único"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{timestamp}.xlsx"

    def cleanup_temp_files(self):
        """Limpia archivos temporales generados"""
        cleaned_count = 0
        for temp_file in self.temp_files:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
                    cleaned_count += 1
            except Exception as e:
                print(f"Error eliminando archivo temporal {temp_file}: {e}")
        self.temp_files = []
        return cleaned_count

    def get_temp_files(self):
        """Obtiene lista de archivos temporales"""
        return self.temp_files.copy()


class ReportService:
    """Servicio principal de generación de reportes"""

    def __init__(self, registry_manager, email_tab=None):
        self.registry_manager = registry_manager
        self.email_tab = email_tab
        self.report_generator = ReportGenerator()

    def get_records_for_report(self, report_type, filter_info=None):
        """Obtiene registros según el tipo de reporte seleccionado"""
        if report_type == ReportTypes.CURRENT_FILTERED and filter_info:
            return self.registry_manager.get_filtered_records(
                date_from=filter_info.get('date_from'),
                date_to=filter_info.get('date_to'),
                status_filter=filter_info.get('status'),
                user_filter=filter_info.get('user'),
                profile_filter=filter_info.get('profile')
            )
        else:
            return self.registry_manager.get_records_by_type(report_type)

    def validate_report_request(self, report_type, filter_info=None):
        """Valida solicitud de reporte antes de generarlo"""
        records = self.get_records_for_report(report_type, filter_info)

        if not records:
            return False, "No hay registros para generar el reporte según los criterios seleccionados"

        if len(records) > 10000:
            return False, "Demasiados registros (>10,000). Aplique filtros para reducir el conjunto de datos"

        # Verificar openpyxl directamente en lugar de usar el método del generador
        try:
            import openpyxl
        except ImportError:
            return False, "openpyxl no está instalado. Instale con: pip install openpyxl"

        return True, f"Reporte válido con {len(records)} registros"

    def export_to_excel(self, report_type, filter_info=None, filename=None):
        """Exporta reporte a archivo Excel"""
        is_valid, message = self.validate_report_request(report_type, filter_info)
        if not is_valid:
            return None, False, message

        try:
            records = self.get_records_for_report(report_type, filter_info)

            if filename is None:
                filename = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx")],
                    initialfile=f"reporte_ejecuciones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                )

                if not filename:
                    return None, False, "Exportación cancelada por el usuario"

            if report_type == ReportTypes.CURRENT_FILTERED and filter_info:
                filter_summary = self._create_filter_summary(filter_info)
                report_title = f"Reporte de Ejecuciones - {filter_summary}"
            else:
                report_title = f"Reporte de Ejecuciones - {report_type}"

            result_filename, success, result_message = self.report_generator.export_to_excel(
                records, filename, report_title
            )

            return result_filename, success, result_message

        except Exception as e:
            return None, False, f"Error inesperado exportando reporte: {str(e)}"

    def _create_filter_summary(self, filter_info):
        """Crea resumen de filtros para título del reporte"""
        parts = []
        if filter_info.get('date_from'):
            parts.append(f"Desde {filter_info['date_from']}")
        if filter_info.get('date_to'):
            parts.append(f"Hasta {filter_info['date_to']}")
        if filter_info.get('status') and filter_info['status'] != "Todos":
            parts.append(f"Estado {filter_info['status']}")
        if filter_info.get('user') and filter_info['user'] != "Todos":
            parts.append(f"Usuario {filter_info['user']}")
        if filter_info.get('profile') and filter_info['profile'] != "Todos":
            parts.append(f"Perfil {filter_info['profile']}")

        return " | ".join(parts) if parts else "Sin filtros"


class EmailReportService:
    """Servicio especializado en envío de reportes por email"""

    def __init__(self, report_service, email_tab):
        self.report_service = report_service
        self.email_tab = email_tab

    def validate_email_setup(self):
        """Valida que el email esté configurado correctamente"""
        if not self.email_tab:
            return False, "Sistema de email no disponible"

        if not self.email_tab.is_email_configured():
            return False, "Email no configurado. Configure el email en la pestaña 'Email'"

        return True, "Email configurado correctamente"

    def send_report_async(self, report_type, filter_info=None, custom_title=None, callback=None):
        """Envía reporte por email de forma asíncrona"""

        def send_thread():
            try:
                success, message = self.send_report_sync(report_type, filter_info, custom_title)
                if callback:
                    callback(success, message)
            except Exception as e:
                if callback:
                    callback(False, f"Error inesperado: {str(e)}")

        threading.Thread(target=send_thread, daemon=True).start()

    def send_report_sync(self, report_type, filter_info=None, custom_title=None):
        """Envía reporte por email de forma síncrona"""
        email_valid, email_message = self.validate_email_setup()
        if not email_valid:
            return False, email_message

        report_valid, report_message = self.report_service.validate_report_request(report_type, filter_info)
        if not report_valid:
            return False, report_message

        try:
            records = self.report_service.get_records_for_report(report_type, filter_info)
            temp_filename = self.report_service.report_generator.create_temp_filename("email_reporte")

            if custom_title:
                report_title = custom_title
            elif report_type == ReportTypes.CURRENT_FILTERED and filter_info:
                filter_summary = self.report_service._create_filter_summary(filter_info)
                report_title = f"Reporte de Ejecuciones - {filter_summary}"
            else:
                report_title = f"Reporte de Ejecuciones - {report_type}"

            filename, success, message = self.report_service.report_generator.export_to_excel(
                records, temp_filename, report_title
            )

            if not success:
                return False, f"Error generando Excel: {message}"

            subject, body = self._prepare_email_content(report_title, records, report_type)
            email_success, email_message = self._send_email_with_attachment(subject, body, filename)

            self._cleanup_temp_file(filename)
            return email_success, email_message

        except Exception as e:
            return False, f"Error enviando reporte: {str(e)}"

    def _prepare_email_content(self, report_title, records, report_type):
        """Prepara contenido del email"""
        subject = f"Syncro Bot - {report_title} - {datetime.now().strftime('%d/%m/%Y')}"
        stats = self._calculate_report_statistics(records)

        body = f"""Estimado/a,

Se adjunta el reporte de ejecuciones del sistema Syncro Bot correspondiente a: {report_type}

RESUMEN EJECUTIVO:
Total de Ejecuciones: {stats['total']}
Ejecuciones Exitosas: {stats['successful']}
Ejecuciones Fallidas: {stats['failed']}
Ejecuciones Manuales: {stats['manual']}
Ejecuciones Automáticas: {stats['automatic']}
Tasa de Éxito: {stats['success_rate']:.1f}%

PERÍODO DEL REPORTE:
Fecha de Generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}
Tipo de Reporte: {report_type}

El archivo adjunto contiene el detalle completo de todas las ejecuciones incluidas en este reporte.

Saludos cordiales,
Sistema Syncro Bot"""

        return subject, body

    def _calculate_report_statistics(self, records):
        """Calcula estadísticas para el reporte"""
        total = len(records)
        successful = len([r for r in records if r['estado'] == 'Exitoso'])
        failed = len([r for r in records if r['estado'] == 'Fallido'])
        manual = len([r for r in records if r['usuario'] == 'Usuario'])
        automatic = len([r for r in records if r['usuario'] == 'Sistema'])

        return {
            'total': total,
            'successful': successful,
            'failed': failed,
            'manual': manual,
            'automatic': automatic,
            'success_rate': (successful / total * 100) if total > 0 else 0
        }

    def _send_email_with_attachment(self, subject, body, attachment_path):
        """Envía email con archivo adjunto"""
        try:
            if not os.path.exists(attachment_path):
                return False, f"El archivo {attachment_path} no existe"

            return self.email_tab.send_email(
                subject=subject,
                body=body,
                attachments=[attachment_path]
            )
        except Exception as e:
            return False, f"Error enviando email: {str(e)}"

    def _cleanup_temp_file(self, filename):
        """Limpia archivo temporal"""
        try:
            if os.path.exists(filename):
                os.remove(filename)
        except Exception as e:
            print(f"Warning: No se pudo limpiar archivo temporal {filename}: {e}")


class AutomaticReportService:
    """Servicio para reportes automáticos desde perfiles"""

    def __init__(self, email_report_service):
        self.email_report_service = email_report_service

    def send_profile_report(self, profile_config, execution_record=None):
        """Envía reporte según configuración del perfil"""
        if not profile_config.get('send_report', False):
            return True, "Reporte no configurado para este perfil"

        try:
            report_frequency = profile_config.get('report_frequency', 'Después de cada ejecución')
            report_type = profile_config.get('report_type', 'Últimos 7 días')
            profile_name = profile_config.get('name', 'Perfil desconocido')

            if report_frequency == "Después de cada ejecución":
                actual_report_type = ReportTypes.LAST_7_DAYS if report_type == "Solo Ejecuciones del Perfil" else report_type
                custom_title = f"Reporte Automático - Perfil '{profile_name}'"

                return self.email_report_service.send_report_sync(
                    report_type=actual_report_type,
                    custom_title=custom_title
                )
            else:
                return False, f"Frecuencia '{report_frequency}' no implementada aún"

        except Exception as e:
            return False, f"Error enviando reporte automático: {str(e)}"


class ReportManager:
    """Gestor principal que coordina todos los servicios de reportes"""

    def __init__(self, registry_manager, email_tab=None):
        self.report_service = ReportService(registry_manager, email_tab)

        if email_tab:
            self.email_report_service = EmailReportService(self.report_service, email_tab)
            self.automatic_report_service = AutomaticReportService(self.email_report_service)
        else:
            self.email_report_service = None
            self.automatic_report_service = None

    def export_to_excel(self, report_type, filter_info=None, filename=None):
        """Exporta reporte a Excel"""
        return self.report_service.export_to_excel(report_type, filter_info, filename)

    def send_report_by_email(self, report_type, filter_info=None, custom_title=None, async_mode=True, callback=None):
        """Envía reporte por email"""
        if not self.email_report_service:
            return False, "Servicio de email no disponible"

        if async_mode:
            self.email_report_service.send_report_async(report_type, filter_info, custom_title, callback)
            return True, "Enviando reporte en segundo plano..."
        else:
            return self.email_report_service.send_report_sync(report_type, filter_info, custom_title)

    def send_automatic_report(self, profile_config, execution_record=None):
        """Envía reporte automático desde perfil"""
        if not self.automatic_report_service:
            return False, "Servicio de reportes automáticos no disponible"

        return self.automatic_report_service.send_profile_report(profile_config, execution_record)

    def get_available_report_types(self):
        """Obtiene tipos de reportes disponibles"""
        return ReportTypes.get_all_types()

    def cleanup_resources(self):
        """Limpia recursos del sistema de reportes"""
        cleaned_count = self.report_service.report_generator.cleanup_temp_files()
        return f"Se limpiaron {cleaned_count} archivos temporales"